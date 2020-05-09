#!/usr/bin/env python3

import csv
import sys

from copy import deepcopy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell

if __name__ == '__main__':
    xlsx = Path(sys.argv[1])
    csvs = xlsx.parent / '_csv_'
    csvs.mkdir(parents=True, exist_ok=True)

    for f in xlsx.glob('*.xlsx'):
        if f.name.startswith('~$'): continue
        print(f)

        wb = load_workbook(f)
        ws = wb.worksheets[0]
        wb_cached = load_workbook(f, data_only=True)
        ws_cached = wb_cached.worksheets[0]

        data = [['' for _ in range(ws.max_column)] for _ in range(ws.max_row)]
        cached_formulae = {}
        for row in ws.rows:
            for cell in row:  # type: Cell
                r = cell.row - 1
                c = cell.column - 1
                dt = cell.data_type
                if dt == 's':
                    data[r][c] = '<str>'
                elif dt == 'n':
                    data[r][c] = cell.value
                elif dt == 'f':
                    data[r][c] = cell.value
                    cached_formulae[(r, c)] = ws_cached.cell(cell.row, cell.column).value
                elif dt == 'e':
                    data[r][c] = '<error>'
                elif dt == 'd' or cell.is_date:
                    date: datetime = cell.value
                    data[r][c] = str(date)
                else:
                    data[r][c] = f'<unkdt:{dt}>'

        for m in ws.merged_cells:  # type: CellRange
            tl = ws.cell(m.min_row, m.min_col)
            for row in ws.iter_rows(m.min_row, m.max_row, m.min_col, m.max_col):
                for cell in row:
                    if cell == tl: continue
                    r = cell.row - 1
                    c = cell.column - 1
                    data[r][c] = f'<m:{tl.coordinate}>'

        data_cached = deepcopy(data)
        for (r, c), value in cached_formulae.items():
            data_cached[r][c] = value

        with open(csvs / f.with_suffix('.csv').name, 'w', newline='', encoding='utf-8') as out:
            csv.writer(out, dialect='excel').writerows(data)
        with open(csvs / f.with_suffix('.eval.csv').name, 'w', newline='', encoding='utf-8') as out:
            csv.writer(out, dialect='excel').writerows(data_cached)